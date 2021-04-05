import xlsxwriter
from datetime import date, datetime
from math import floor
import openpyxl
from db import delete_all_data, store_all_data, get_alerts_by_sla, get_alerts_by_state, get_alerts_by_priority, \
    get_alerts_by_type, get_alerts_by_severity, get_alerts_by_owner

date_s = date.today()
date_s = floor(date_s.day / 10) * 10
if date_s < 10:
    week = 'Week 3'
elif 10 <= date_s < 20:
    week = 'Week 1'
else:
    week = 'Week 2'
name_dest = 'Monthly KPI Report ' + week + '.xlsx'

file, file_flag = 0, 0
while file_flag == 0:
    try:
        file = input('Enter the path or name of you file: ')
        # file = 'Monthly KPI Data Week 1.xlsx'
        # if file.__contains__('\\'):
        #     file = file.replace('\\','\\\\')
        book = openpyxl.load_workbook(filename=file)
        file_flag = 1
    except FileNotFoundError:
        print('File Not Found.')
    except OSError:
        print('Enter a Valid Path or Name.')
book = openpyxl.load_workbook(file)
sheet = book['Data']
header_flag = True
i = 0
header = []
while header_flag:
    i += 1
    if sheet.cell(1, i).value:
        temp = sheet.cell(1, i).value
        if temp[0] == " ":
            temp = temp[1:]
        header.append(temp.lower())

    else:
        header_flag = False

data_flag = True
data = []
k = 2
while data_flag:
    if sheet.cell(k, 1).value:
        temp = {}
        for j in range(1, i):
            temp.update({header[j - 1]: sheet.cell(k, j).value})
        data.append(temp)
    else:
        data_flag = False
    k += 1

book = xlsxwriter.Workbook(name_dest)
try:
    store_all_data(data)
    sheet0 = book.add_worksheet(name='Data')

    formats = {
        'number': book.add_format({'num_format': '0', 'align': 'center', 'border': 1}),
        'border': book.add_format({'align': 'center', 'border': 1}),
        'heading': book.add_format({'bg_color': '#000000', 'align': 'center', 'font_color': '#FFFFFF'}),
        'date': book.add_format({'num_format': 'dd/mmm/yy hh:mm', 'border': 1})
    }
    sheet = book.add_worksheet(name='State')
    sheet.hide_gridlines(2)

    result = get_alerts_by_state()
    sheet.write(0, 0, 'States', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    sheet = book.add_worksheet(name='Priority')
    sheet.hide_gridlines(2)

    result = get_alerts_by_priority()
    sheet.write(0, 0, 'Priority', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    sheet = book.add_worksheet(name='SLA')
    sheet.hide_gridlines(2)

    result = get_alerts_by_sla()
    sheet.write(0, 0, 'SLA Labels', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    sheet = book.add_worksheet(name='Type')
    sheet.hide_gridlines(2)

    result = get_alerts_by_type()
    sheet.write(0, 0, 'Type Lables', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    sheet = book.add_worksheet(name='Severity')
    sheet.hide_gridlines(2)

    result = get_alerts_by_severity()
    sheet.write(0, 0, 'Type Lables', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    agents = {
        'noc': 'NOC',
        'nouman': 'Nouman',
        'abdullah': 'Abdullah',
        'arslan': 'Arslan',
        'mohsin': 'Mohsin',
        'samuel': 'Samuel',
        'latifa': 'Queen',
        'prince': 'Prince',
        'essien': 'Essien',
        'dine': 'Dine',
        'zeeshan': 'Zeeshan',
        'duchet': 'Duchet',
        'omer': 'Omer',
        'salahuddin': 'Salahuddin',
        'yapo': 'Yapo',
        'nadeem': 'Nadeem',
        'ahmed': 'Ahmed',
        'ayyaz': 'Ayyaz',
        'ahsan': 'Ahsan',
        'michael': 'Michael',
        'peter': 'Peter',
        'omair': 'Omair',
        'barun': 'Barun',
        'emmanuel': 'Emmanuel',
        'amarendra': 'Amarendra',
        'debadrita': 'Debadrita',
    }

    sheet = book.add_worksheet(name='Agent Owner')
    sheet.hide_gridlines(2)

    result = list(get_alerts_by_owner())
    sheet.write(0, 0, 'Type Lables', formats['heading'])
    sheet.write(0, 1, 'Count', formats['heading'])
    row = 1
    for key in result:
        key = list(key)
        for key1 in agents:
            if key[0].lower().__contains__(key1):
                key[0] = agents[key1]
        sheet.write(row, 0, key[0], formats['border'])
        sheet.write(row, 1, key[1], formats['border'])
        row += 1
    sheet.write(row, 0, 'Total', formats['heading'])
    sheet.write(row, 1, '=SUM($B$2:$B$' + str(row) + ')', formats['heading'])

    sheet0.hide_gridlines(2)
    col = 0
    header = {}
    for key in list(data[0].keys()):
        temp = ''
        if key == 'sla':
            temp = 'SLA'
        elif key == 'firstresponse':
            temp = 'First Response'
        elif key == 'solutiontime':
            temp = 'Solution Time'
        elif key == 'agent/owner':
            temp = 'Agent/Owner'
        else:
            for key1 in key.split():
                temp += ' ' + key1[0].upper() + key1[1:]
        if key.lower().__contains__('queue'):
            sheet0.write(0, len(list(data[0].keys())) - 1, temp, formats['heading'])
            header.update({key: len(list(data[0].keys())) - 1})
        else:
            sheet0.write(0, col, temp, formats['heading'])
            header.update({key: col})
            col += 1

    row = 1
    for key in data:
        # if not delete.__contains__(key):
            for key1 in header:
                try:
                    if type(key[key1]) is datetime:
                        sheet0.write(row, header[key1], key[key1], formats['date'])
                    else:
                        sheet0.write(row, header[key1], int(key[key1]), formats['number'])
                except:
                    sheet0.write(row, header[key1], key[key1], formats['border'])
            # print(key)
            row += 1


finally:
    delete_all_data()
    book.close()
input('Enter to Close.')
